import openpyxl
import sqlite3 as sq
import plotly.graph_objs as go

# wookbook = openpyxl.load_workbook("C:\MYprog\Project\loaded\Students.xlsx")
# worksheet, dataup = wookbook.active, []
# for i in range(0, worksheet.max_row):
#     c = []
#     for col in worksheet.iter_cols(1, worksheet.max_column):
#         c.append(col[i].value)
#     dataup.append(c)
# for i in range(0, worksheet.max_row):
#     for col in worksheet.iter_cols(1, worksheet.max_column):
#         print(col[i].value, end="\t")
#     print('')

connection = sq.connect("C:\MYprog\Project\static\DataS\ProjectSite.db")
cursor = connection.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS Studentsvisit (data TEXT,groups TEXT,surname TEXT,name TEXT,patronymic TEXT,firstentrance TEXT,lastexit TEXT,attend TEXT,status TEXT)''')
data = cursor.execute("SELECT data, groups, surname, name FROM Studentsvisit")
data = [*data]
c = [tuple(i) for i in dataup[1::]]
# for i in c:
#     cursor.execute('''INSERT INTO Studentsvisit (data, groups, surname, name, patronymic, firstentrance, lastexit, attend, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', i)
# print(*data)
data = set(tuple([i[0], i[1], i[2], i[3], i[4], i[5]]) for i in cursor.execute("SELECT data, groups, surname, name, patronymic, attend FROM Studentsvisit"))
datatms = set(tuple([i[0], i[1], i[2], i[3]]) for i in cursor.execute("SELECT groups, surname, name, patronymic FROM Studentsvisit"))
times = len(set(cursor.execute("SELECT data FROM Studentsvisit")))
wordd, datagr = {}, []
for i in datatms:
    c = 0
    if i not in wordd.keys():
        wordd[i] = 0
    for j in data:
        if i == tuple([j[1], j[2], j[3], j[4]]) and j[5] == "да":
            c = c + 1
    wordd[i] = round(c , 2)
for i in wordd.keys():
    datagr.append((f"{i[1]} {i[2][0].upper()}. {i[3][0].upper()}.", wordd[i]))
lengr = len(datagr)
c = 255
colorbr = []
for i in range(lengr):
    colorbr.append([0, 0, round(c)])
    c -= 255 / lengr
datagr.sort(key=lambda x: x[1], reverse=True)
datag = [[], []]
for i in range(lengr):
    datag[0].append(datagr[i][0])
    datag[1].append(datagr[i][1])
const_max = times
lr = lengr
arg = 8
fig = go.Figure()
for step in range(lengr - arg):
    if step == 0:
        fig.add_trace(go.Bar(visible=False,y=datag[1] ))
    else:
        fig.add_trace(go.Bar(visible=False,x=datag[0][step-1:step+arg-1] + ["void"],y=datag[1][step-1:step+arg-1] + [const_max]))
fig.data[0].visible = True
go.Bar()
steps = []
for i in range(len(fig.data)):
    step = dict(method="update",args=[{"visible": [False] * len(fig.data)}],)
    step["args"][0]["visible"][i] = True  # Toggle i'th trace to "visible"
    steps.append(step)
sliders = [dict(active=0,currentvalue={"prefix": "Frequency: "},pad={"t": 11},steps=steps)]
fig.update_layout(sliders=sliders)
fig.show()
connection.commit()
connection.close()
