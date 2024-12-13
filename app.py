import openpyxl
import sqlite3 as sq

wookbook = openpyxl.load_workbook("sales.xlsx")
worksheet, dataup = wookbook.active, []
for i in range(0, worksheet.max_row):
    c = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        c.append(col[i].value)
    dataup.append(c)
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t")
    print('')

connection = sq.connect("Students.db")
cursor = connection.cursor()
cursor.execute('''
CREATE TABLE IF NOT EXISTS Studentsvisit (
data TEXT,
groups TEXT,
surname TEXT,
name TEXT,
patronymic TEXT,
firstentrance TEXT,
lastexit TEXT,
attend TEXT,
status TEXT
)
''')
c = [tuple(i) for i in dataup[1::]]
for i in c:
    cursor.execute('''INSERT INTO Studentsvisit (data, groups, surname, name, patronymic, firstentrance, lastexit, attend, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', i)

connection.commit()
connection.close()

import plotly.graph_objects as go
import numpy as np

# Create figure
xb = ["f", 'fas', 'fw', 'wev', 'qfsa', 'afsaw', 'rtyrtg', 'trshrh', 'trhdht', 'dsht', 'sthrstssz']
yb = [1, 4, 5, 4, 4, 3, 2, 1, 2, 4, 5]
fig = go.Figure()

# Add traces, one for each slider step
for step in range(2):
    print(step)
    fig.add_trace(
        go.Bar(
            visible=False,
            x=xb[step:step+5],
            y=yb[step:step+5]))

# Make 10th trace visible
fig.data[0].visible = True

# Create and add slider
steps = []
for i in range(len(fig.data)):
    step = dict(
        method="update",
        args=[{"visible": [False] * len(fig.data)},
              {"title": "Slider switched to step: " + str(i)}],  # layout attribute
    )
    step["args"][0]["visible"][i] = True  # Toggle i'th trace to "visible"
    steps.append(step)

sliders = [dict(
    active=0,
    currentvalue={"prefix": "Frequency: "},
    pad={"t": 11},
    steps=steps
)]

fig.update_layout(sliders=sliders)

fig.show()
