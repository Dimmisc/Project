from openpyxl import load_workbook
from datetime import datetime as dt
from datetime import timedelta

from database.site_data import Students, Grades
from database.data import Visitings
from sqlalchemy.orm import Session




def extand_xlsx_file(db_sess, file_href) -> str:
    print(f"Site start add data from file: {file_href}")
    wookbook = load_workbook(file_href)

    worksheet, ExcelData = wookbook.active, []

    """ Next algoritm changed format from xlsx worbook to python list in rows """
    # elements of row:(1-date; 2-grade; 3-surname; 4-name; 5-patronymic; 6-first enter; 7-last exit; 8-attnding studeent; 9-status)

    for row in worksheet.iter_rows(2, worksheet.max_row):
        c = []
        for i in range(0, worksheet.max_column):
            if i == 0:
                c.append((str(row[i].value).split())[0])
            else:
                c.append(str(row[i].value))
        ExcelData.append(c)

    if db_sess == None:
        return False
    else:
        print(db_sess)
    
    for column in ExcelData:
        grade = db_sess.query(Grades).filter_by(grade=column[1]).first()
        if grade:
            student = db_sess.query(Students).filter_by(grade_name=column[1], 
                                                        surname=column[2], 
                                                        name=column[3], 
                                                        patronymic=column[4]).first()

            visitday = str(column[0]).split('-')
            visitday = dt(int(visitday[0]), int(visitday[1]), int(visitday[2]))

            new_visit = Visitings(date=column[0],
                                  weekDay=visitday.weekday(),
                                  grade=column[1],
                                  surname=column[2],
                                  name=column[3],
                                  patronymic=column[4],
                                  firstEnter=column[5],
                                  lastExit=column[6],
                                  attended=True if column[7] == 'Да' else False,
                                  status=column[8]
                                  )
            
            print(new_visit.grade)

            if student:
                new_visit.student = student
                db_sess.add(new_visit)
            elif column[8] == "Обучающиеся":
                new_student = Students(grade_name=column[1],
                                       surname=column[2],
                                       name=column[3],
                                       patronymic=column[4],
                                       status=column[8],
                                       )
                new_student.Grade = grade
                db_sess.add(new_student)
                new_visit.student = new_student
                db_sess.add(new_visit)
        else:
            new_grade = Grades(grade=column[1])
            new_student = Students(grade_name=column[1],
                                   surname=column[2],
                                   name=column[3],
                                   patronymic=column[4],
                                   status=column[8],
                                   )
            

            visitday = str(column[0]).split('-')
            visitday = dt(int(visitday[0]), int(visitday[1]), int(visitday[2]))

            new_visit = Visitings(date=column[0],
                                  weekDay=visitday.weekday(),
                                  grade=column[1],
                                  surname=column[2],
                                  name=column[3],
                                  patronymic=column[4],
                                  firstEnter=column[5],
                                  lastExit=column[6],
                                  attended=True if column[7] == 'Да' else False,
                                  status=column[8]
                                  )
            db_sess.add(new_grade)
            new_student.Grade = new_grade
            db_sess.add(new_student)
            new_visit.student = new_student
            db_sess.add(new_visit)
    db_sess.commit()

    print(f"Site ended add data from file: {file_href}")

    return str(dt.today()).split()[0]


def GetDataStudents(db_sess) -> list:
    data = [[], []]
    students = db_sess.query(Students).all()

    for student in students:
        student_abbreviation = f'<a href="/provide_student_visiting/{student.id}">{str(student.surname)[0].upper()}. {str(student.name)[0].upper()}. {str(student.patronymic)[0].upper()}</a>'
        student_visit = 0
        for visit in student.Visits:
            if visit.attended == True:
                print(visit.attended)
                student_visit += 1
        data[0].append(student_abbreviation)
        data[1].append(student_visit)

    return data


def GetGradesData(db_sess) -> list:
    grades = db_sess.query(Grades).all()
    data = [[], []]
    if grades:
        for grade in grades:
            grade_label = f"<a href='{grade.id}'>{grade.grade}</a>"
            maxvisit = 0
            visitings = 0
            for student in grade.students:
                for visit in student.Visits:
                    if visit.attended == True:
                        visitings += 1
                    maxvisit += 1
            data[0].append(grade_label)
            data[1].append(visitings)
    else:
        return False
    data.append(maxvisit)
    """ Function returned data for graphic and max avalible visitng of students in one grade at the end """
    return data


def CheckDateVisitings(href_to_file) -> str:
    xlsx_file = load_workbook(href_to_file)
    xlsx_file = xlsx_file.active
    date = xlsx_file.iter_rows(1, 2)[1]
    return date


def GetDataStudent(db_sess, id_student) -> list:
    data = [[],[]]
    student = db_sess.query(Students).filter_by(id=id_student).first()
    visitings = student.Visits

    min_day = min([(visit.date, visit.weekDay) for visit in visitings], key=lambda x: x[0])
    max_day = max([(visit.date, visit.weekDay) for visit in visitings], key=lambda x: x[0])

    miday = min_day[0].split('-')
    maday = max_day[0].split('-')

    min_day = [dt(int(miday[0]), int(miday[1]), int(miday[2])), min_day[1]]
    max_day = [dt(int(maday[0]), int(maday[1]), int(maday[2])), max_day[1]]

    min_day[0] -= timedelta(days=int(min_day[1]))
    max_day[0] += timedelta(days=7 - int(max_day[1]))

    min_day, max_day = min_day[0], max_day[0]
    one_day_delta = timedelta(days=1)
    data[0].extend(['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'])
    for i in range((max_day - min_day).days):
        min_day += one_day_delta
        str_TIME = str(min_day).split()[0]

        print(str_TIME)
        
        visit = [*filter(lambda element: str_TIME == element.date, visitings)]
        if visit:
            data[1].append([visit[0].firstEnter, visit[0].attended])
        else:
            data[1].append(['None data', False])
    return data